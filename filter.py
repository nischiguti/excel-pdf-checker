import os
from openpyxl import load_workbook
from PyPDF2 import PdfReader

# Define the root directory containing all the equipment folders
root_dir = '/home/freguez/Documents/equipamentos/'  # Replace with your actual root directory path
output_file = '/home/freguez/Documents/equipamentos/dados.txt'  # Replace with your desired output file path


# Function to extract text from a PDF
def extract_text_from_pdf(pdf_path):
    reader = PdfReader(pdf_path)
    text = ""
    for page in reader.pages:
        text += page.extract_text()
    return text

# Function to check if the numbers are present in the PDF text
def check_pdf_for_numbers(pdf_path, equipment_number, serial_number):
    text = extract_text_from_pdf(pdf_path)
    equipment_found = equipment_number in text if equipment_number else False
    serial_found = serial_number in text if serial_number else False
    return equipment_found, serial_found

def find_value_next_to_label(sheet, label):
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and label.lower() in str(cell.value).lower():
                # Start searching from the next column and keep moving right until a non-empty cell is found
                current_column = cell.column + 1
                while sheet.cell(row=cell.row, column=current_column).value is None:
                    current_column += 1
                    # Break if out of bounds (highly unlikely in well-formed Excel sheets)
                    if current_column > sheet.max_column:
                        break
                value = sheet.cell(row=cell.row, column=current_column).value
                return str(value).strip() if value else None
    return None


# Open the output file for writing
with open(output_file, 'w') as log_file:
    # Loop through each equipment folder
    for folder_name in os.listdir(root_dir):
        folder_path = os.path.join(root_dir, folder_name)
        
        if os.path.isdir(folder_path) and folder_name[:4].isdigit():
            equipment_number = folder_name[:4]  # Extract the equipment number from the first 4 characters of the folder name
            
            # Find the Excel file with 'Ficha' in its name
            excel_file = None
            for file_name in os.listdir(folder_path):
                if file_name.lower().endswith('.xlsx') and 'ficha' in file_name.lower():
                    excel_file = os.path.join(folder_path, file_name)
                    break
            
            if excel_file:
                # Load the Excel file and extract serial number
                workbook = load_workbook(excel_file, data_only=True)
                sheet = workbook.active
                serial_number = find_value_next_to_label(sheet, "(Serial Number)")

                # Check if the serial number was found
                if serial_number:
                    # Check the subfolder containing 'calibração' for PDFs
                    for subfolder_name in os.listdir(folder_path):
                        subfolder_path = os.path.join(folder_path, subfolder_name)
                        
                        if os.path.isdir(subfolder_path) and 'calibração' in subfolder_name.lower():
                            for pdf_name in os.listdir(subfolder_path):
                                if pdf_name.endswith('.pdf'):
                                    pdf_path = os.path.join(subfolder_path, pdf_name)
                                    equipment_found, serial_found = check_pdf_for_numbers(pdf_path, equipment_number, serial_number)
                                    
                                    log_file.write(f"Folder: {folder_name}\n")
                                    log_file.write(f"Extracted Equipment Number: {equipment_number}\n")
                                    log_file.write(f"Excel Serial Number: {serial_number}\n")
                                    log_file.write(f"PDF: {pdf_name}\n")
                                    
                                    if equipment_found and serial_found:
                                        log_file.write(f"Contains both equipment number and serial number.\n")
                                    elif equipment_found:
                                        log_file.write(f"Contains equipment number but not serial number.\n")
                                    elif serial_found:
                                        log_file.write(f"Contains serial number but not equipment number.\n")
                                    else:
                                        log_file.write(f"Does not contain equipment number or serial number.\n")
                                    
                                    log_file.write("\n")
                else:
                    log_file.write(f"Folder: {folder_name}\n")
                    log_file.write("Error: Could not find serial number in the Excel file.\n\n")

print(f"PDF checking complete. Results are saved in {output_file}.")
