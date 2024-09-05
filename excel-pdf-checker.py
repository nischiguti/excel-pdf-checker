import os
from openpyxl import load_workbook
from PyPDF2 import PdfReader

# Define the root directory containing all the equipment folders
root_dir = '/home/freguez/Documents/equipamentos/'  # Replace with your actual root directory path

# Define output files for each category
output_files = {
    "Contains both equipment number and serial number.": '/home/freguez/Documents/equipamentos/contains_both.txt',
    "Contains equipment number but not serial number.": '/home/freguez/Documents/equipamentos/contains_equipment_only.txt',
    "Contains serial number but not equipment number.": '/home/freguez/Documents/equipamentos/contains_serial_only.txt',
    "Does not contain equipment number or serial number.": '/home/freguez/Documents/equipamentos/contains_none.txt',
    "Error: Could not find serial number in the Excel file.": '/home/freguez/Documents/equipamentos/errors.txt'
}

def extract_text_from_pdf(pdf_path):
    """Extract text from a PDF file."""
    reader = PdfReader(pdf_path)
    text = ""
    for page in reader.pages:
        text += page.extract_text()
    return text

def check_pdf_for_numbers(pdf_path, equipment_number, serial_number):
    """Check if the PDF contains the equipment number and serial number."""
    text = extract_text_from_pdf(pdf_path)
    equipment_found = equipment_number in text if equipment_number else False
    serial_found = serial_number in text if serial_number else False
    return equipment_found, serial_found

def find_value_next_to_label(sheet, label):
    """Find the value next to a label in an Excel sheet."""
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and label.lower() in str(cell.value).lower():
                # Start searching from the next column and keep moving right until a non-empty cell is found
                current_column = cell.column + 1
                while sheet.cell(row=cell.row, column=current_column).value is None:
                    current_column += 1
                    if current_column > sheet.max_column:
                        break
                value = sheet.cell(row=cell.row, column=current_column).value
                return str(value).strip() if value else None
    return None

def filter_by_category(input_file, output_files):
    """Filter blocks in the input file into categories and write to corresponding output files."""
    output_handlers = {category: open(file, 'w') for category, file in output_files.items()}
    
    with open(input_file, 'r') as infile:
        current_block = []
        
        for line in infile:
            if line.strip() == "":
                if current_block:
                    block_text = "".join(current_block).strip()
                    last_line = block_text.splitlines()[-1].strip()
                    found = False
                    for category in output_files.keys():
                        if category in last_line:
                            output_handlers[category].write(block_text + "\n\n")
                            found = True
                            break
                    
                    if not found:
                        output_handlers["Error: Could not find equipment number or serial number in the Excel file."].write(block_text + "\n\n")
                
                current_block = []
            else:
                current_block.append(line)
        
        if current_block:
            block_text = "".join(current_block).strip()
            last_line = block_text.splitlines()[-1].strip()
            found = False
            for category in output_files.keys():
                if category in last_line:
                    output_handlers[category].write(block_text + "\n\n")
                    found = True
                    break
            
            if not found:
                output_handlers["Error: Could not find equipment number or serial number in the Excel file."].write(block_text + "\n\n")
    
    for handler in output_handlers.values():
        handler.close()

def main():
    # Prepare output file handlers
    output_handlers = {category: open(file, 'w') for category, file in output_files.items()}

    try:
        # Loop through each equipment folder
        for folder_name in os.listdir(root_dir):
            folder_path = os.path.join(root_dir, folder_name)
            
            if os.path.isdir(folder_path) and folder_name[:4].isdigit():
                equipment_number = folder_name[:4]  # Extract the equipment number from the first 4 characters of the folder name
                
                # Find the Excel file with 'Ficha' in its name
                excel_file = None
                for file_name in os.listdir(folder_path):
                    if file_name.lower().endswith('.xlsx') and 'ficha' in file_name.lower():
                        excel_file = os.path.join(folder_path, file_name)
                        break
                
                if excel_file:
                    # Load the Excel file and extract serial number
                    workbook = load_workbook(excel_file, data_only=True)
                    sheet = workbook.active
                    serial_number = find_value_next_to_label(sheet, "(Serial Number)")

                    # Check if the serial number was found
                    if serial_number:
                        # Check the subfolder containing 'calibração' for PDFs
                        for subfolder_name in os.listdir(folder_path):
                            subfolder_path = os.path.join(folder_path, subfolder_name)
                            
                            if os.path.isdir(subfolder_path) and 'calibração' in subfolder_name.lower():
                                for pdf_name in os.listdir(subfolder_path):
                                    if pdf_name.endswith('.pdf'):
                                        pdf_path = os.path.join(subfolder_path, pdf_name)
                                        equipment_found, serial_found = check_pdf_for_numbers(pdf_path, equipment_number, serial_number)
                                        
                                        log_message = f"Folder: {folder_name}\n"
                                        log_message += f"Extracted Equipment Number: {equipment_number}\n"
                                        log_message += f"Excel Serial Number: {serial_number}\n"
                                        log_message += f"PDF: {pdf_name}\n"
                                        
                                        if equipment_found and serial_found:
                                            log_message += f"Contains both equipment number and serial number.\n"
                                        elif equipment_found:
                                            log_message += f"Contains equipment number but not serial number.\n"
                                        elif serial_found:
                                            log_message += f"Contains serial number but not equipment number.\n"
                                        else:
                                            log_message += f"Does not contain equipment number or serial number.\n"
                                        
                                        log_message += "\n"
                                        output_handlers["Contains both equipment number and serial number." if equipment_found and serial_found else "Contains equipment number but not serial number." if equipment_found else "Contains serial number but not equipment number." if serial_found else "Does not contain equipment number or serial number."].write(log_message)
                    else:
                        output_handlers["Error: Could not find serial number in the Excel file."].write(f"Folder: {folder_name}\nError: Could not find serial number in the Excel file.\n\n")
    finally:
        # Close all output file handlers
        for handler in output_handlers.values():
            handler.close()

    print(f"Processing complete. Results have been saved to the respective files.")

if __name__ == "__main__":
    main()

